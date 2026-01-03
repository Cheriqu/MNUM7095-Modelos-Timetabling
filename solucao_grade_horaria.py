import pandas as pd
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import csv

def solucao_grade_horaria():
    # Carrega a tabela de alocação
    df = pd.read_csv("tabela_alocacao.csv")

    # Substitui nomes das turmas para 1, 2, 3 (como inteiros)
    turma_map = {"c1": 1, "c2": 2, "c3": 3}
    df["Turma"] = df["Turma"].replace(turma_map).astype("Int64")

    # Cria um DataFrame base com Professores como índice e 25 colunas: m1 a m5 repetidos para os 5 dias
    professores = sorted(df['Professor'].unique())
    dias = list(range(5))  # Dias 0 a 4
    periodos = list(range(5))  # Períodos m1 a m5
    colunas = [f"D{d}_m{p+1}" for d in dias for p in periodos]

    # Inicializa a grade vazia
    grade = pd.DataFrame(pd.NA, index=professores, columns=colunas)

    # Preenche a grade com as turmas (como inteiros)
    for _, row in df.iterrows():
        prof = row["Professor"]
        dia = row["Dia"]
        periodo = row["Período"]
        turma = row["Turma"]
        coluna = f"D{dia}_m{periodo+1}"
        grade.loc[prof, coluna] = turma

    # Identifica aulas duplas para cada professor, dia e turma
    aulas_duplas = set()
    for prof in professores:
        for d in dias:
            periodos_prof = []
            turmas_prof = []
            for p in periodos:
                col = f"D{d}_m{p+1}"
                turma = grade.loc[prof, col]
                periodos_prof.append(p)
                turmas_prof.append(turma)
            for p in range(4):  # até o penúltimo período
                t1 = turmas_prof[p]
                t2 = turmas_prof[p+1]
                if pd.notna(t1) and pd.notna(t2) and t1 == t2:
                    aulas_duplas.add((prof, d, p, int(t1)))     # primeira célula da dupla
                    aulas_duplas.add((prof, d, p+1, int(t1)))   # segunda célula da dupla

    # Identifica janelas (períodos vazios entre aulas)
    janelas = set()
    for prof in professores:
        for d in dias:
            ocupados = [p for p in periodos if pd.notna(grade.loc[prof, f"D{d}_m{p+1}"])]
            if len(ocupados) <= 1:
                continue
            for i in range(len(ocupados) - 1):
                for p in range(ocupados[i] + 1, ocupados[i+1]):
                    janelas.add((prof, d, p))

    # Identifica aulas duplas não atendidas (aulas únicas em um dia)
    aulas_unicas = set()
    for prof in professores:
        for d in dias:
            for turma in [1, 2, 3]:
                # Encontra períodos desse professor, dia e turma
                periodos_turma = [
                    p for p in periodos
                    if (pd.notna(grade.loc[prof, f"D{d}_m{p+1}"]) and grade.loc[prof, f"D{d}_m{p+1}"] == turma)
                ]
                if len(periodos_turma) == 1:
                    aulas_unicas.add((prof, d, periodos_turma[0], turma))

    # --- NOVO: Identifica aulas únicas a colorir de amarelo ---
    aulas_unicas_amarelo = set()
    for prof in professores:
        # Coleta todas as aulas únicas desse professor: (dia, periodo, turma)
        unicas_prof = [(d, p, turma) for (pr, d, p, turma) in aulas_unicas if pr == prof]
        # Se o professor tem mais de uma aula única em dias diferentes, todas ficam amarelas
        dias_unicos = set([d for (d, p, turma) in unicas_prof])
        if len(dias_unicos) > 1:
            for (d, p, turma) in unicas_prof:
                aulas_unicas_amarelo.add((prof, d, p, turma))
    # --- Marca todas as aulas únicas não consecutivas no mesmo dia ---
    for prof in professores:
        for d in dias:
            # Coleta todas as aulas únicas desse professor nesse dia
            unicas_dia = [(p, turma) for (pr, dd, p, turma) in aulas_unicas if pr == prof and dd == d]
            if len(unicas_dia) > 1:
                # Ordena por período
                unicas_dia_sorted = sorted(unicas_dia, key=lambda x: x[0])
                # Marca todas as aulas únicas desse dia como amarelas se não forem consecutivas
                for idx in range(len(unicas_dia_sorted) - 1):
                    p1, turma1 = unicas_dia_sorted[idx]
                    p2, turma2 = unicas_dia_sorted[idx + 1]
                    if p2 != p1 + 1:
                        aulas_unicas_amarelo.add((prof, d, p1, turma1))
                        aulas_unicas_amarelo.add((prof, d, p2, turma2))
    # --- NOVO: Marca aulas da MESMA turma no mesmo dia não consecutivas ---
    for prof in professores:
        for d in dias:
            for turma in [1, 2, 3]:
                # Pega todos os períodos dessa turma para esse professor e dia
                periodos_turma = [
                    p for p in periodos
                    if (pd.notna(grade.loc[prof, f"D{d}_m{p+1}"]) and grade.loc[prof, f"D{d}_m{p+1}"] == turma)
                ]
                if len(periodos_turma) > 1:
                    periodos_turma_sorted = sorted(periodos_turma)
                    for idx in range(len(periodos_turma_sorted) - 1):
                        p1 = periodos_turma_sorted[idx]
                        p2 = periodos_turma_sorted[idx + 1]
                        if p2 != p1 + 1:
                            aulas_unicas_amarelo.add((prof, d, p1, turma))
                            aulas_unicas_amarelo.add((prof, d, p2, turma))
    # ----------------------------------------------------------

    # Cores
    cor_azul = "FF00B0F0"    # Azul para janelas
    cor_cinza_fraco = "FFEEEEEE"
    cor_cinza_medio = "FFCCCCCC"
    cor_cinza_forte = "FF888888"
    cor_amarelo = "FFFFFF00" # Amarelo para aulas únicas destacadas
    cor_vazio = "FFFFFFFF"   # Branco para células vazias

    # Cria um objeto ExcelWriter
    writer = pd.ExcelWriter("solucao_grade_horaria.xlsx", engine='openpyxl')
    grade.to_excel(writer, sheet_name='Grade Horária', index_label="Professor")

    # Obtém a planilha do workbook
    workbook = writer.book
    sheet = writer.sheets['Grade Horária']

    # Adiciona linha extra para os dias e ajusta os períodos
    sheet.insert_rows(1)
    # Escreve os dias mesclando as células
    for d in range(5):
        start_col = 2 + d * 5
        end_col = start_col + 4
        cell = sheet.cell(row=1, column=start_col)
        cell.value = f"Dia {d+1}"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    # Escreve os períodos na linha 2
    for d in range(5):
        for p in range(5):
            col = 2 + d * 5 + p
            cell = sheet.cell(row=2, column=col)
            cell.value = f"{p+1}"
            cell.alignment = Alignment(horizontal="center", vertical="center")
    # Ajusta o cabeçalho da coluna de índice
    sheet.cell(row=1, column=1).value = ""
    sheet.cell(row=2, column=1).value = "Professor"

    # Define bordas
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="medium", color="000000")

    # Ajusta largura das colunas (exceto a primeira)
    for col_idx in range(2, len(grade.columns) + 2):
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = 3

    # Aplica as cores, centralização e bordas às células de dados
    for row_idx, professor in enumerate(grade.index):
        for col_idx, col_name in enumerate(grade.columns):
            cell_value = grade.loc[professor, col_name]
            excel_row = row_idx + 3  # +3 por causa das duas linhas de cabeçalho e índice
            excel_col = col_idx + 2

            cell = sheet.cell(row=excel_row, column=excel_col)

            # Descobre dia e período a partir do nome da coluna
            d = col_idx // 5
            p = col_idx % 5

            # Cor de fundo e valor
            if pd.isna(cell_value):
                # Se for janela, pinta de azul
                if (professor, d, p) in janelas:
                    cell.fill = PatternFill(start_color=cor_azul, end_color=cor_azul, fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color=cor_vazio, end_color=cor_vazio, fill_type="solid")
                cell.value = None
            else:
                turma_num = int(cell_value)
                # Aulas duplas: cor de acordo com a turma
                if (professor, d, p, turma_num) in aulas_duplas:
                    if turma_num == 1:
                        cor_cinza = cor_cinza_fraco
                    elif turma_num == 2:
                        cor_cinza = cor_cinza_medio
                    elif turma_num == 3:
                        cor_cinza = cor_cinza_forte
                    else:
                        cor_cinza = cor_vazio
                    cell.fill = PatternFill(start_color=cor_cinza, end_color=cor_cinza, fill_type="solid")
                # NOVO: aulas únicas amarelas
                elif (professor, d, p, turma_num) in aulas_unicas_amarelo:
                    cell.fill = PatternFill(start_color=cor_amarelo, end_color=cor_amarelo, fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color=cor_vazio, end_color=cor_vazio, fill_type="solid")
                cell.value = turma_num  # Armazena como inteiro
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Define bordas para cada célula
            left = thick if excel_col == 2 else thin
            top = thick if excel_row == 3 else thin
            right = thick if (excel_col - 2 + 1) % 5 == 0 and excel_col > 2 else thin
            bottom = thin

            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # Bordas grossas no cabeçalho (primeira e segunda linha)
    for col_idx in range(len(grade.columns) + 1):
        for row in [1, 2]:
            cell = sheet.cell(row=row, column=col_idx + 1)
            left = thick if col_idx == 0 else thin
            right = thick if (col_idx > 0 and col_idx % 5 == 0) else thin
            top = thick
            bottom = thin
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # Bordas grossas na primeira coluna (índice dos professores)
    for row_idx in range(len(grade.index) + 2):
        cell = sheet.cell(row=row_idx + 1, column=1)
        left = thick
        top = thick if row_idx == 0 else thin
        right = thin
        bottom = thin
        cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    # Salva o arquivo Excel
    writer.close()

def calcular_metricas_e_adicionar_excel():
    import json

    df = pd.read_csv("tabela_alocacao.csv")
    turma_map = {"c1": 1, "c2": 2, "c3": 3}
    df["Turma"] = df["Turma"].replace(turma_map).astype("Int64")

    # Calcula aulas duplas realizadas por evento
    duplas_realizadas = {e: 0 for e in df["Evento"].unique()}
    max_duplas_possiveis = {e: 0 for e in df["Evento"].unique()}
    for e in df["Evento"].unique():
        df_e = df[df["Evento"] == e]
        for dia in df_e["Dia"].unique():
            periodos = sorted(df_e[df_e["Dia"] == dia]["Período"].tolist())
            # Conta duplas realizadas
            for i in range(len(periodos) - 1):
                if periodos[i+1] == periodos[i] + 1:
                    duplas_realizadas[e] += 1
            # Máximo de duplas possíveis nesse dia
            max_duplas_possiveis[e] += len(periodos) // 2

    # Calcula aulas duplas não atendidas por evento (máximo possível - realizadas)
    duplas_nao_atendidas = 0
    for e in duplas_realizadas:
        faltam = max(0, max_duplas_possiveis[e] - duplas_realizadas[e])
        duplas_nao_atendidas += faltam

    wb = load_workbook("solucao_grade_horaria.xlsx")
    ws = wb["Grade Horária"]

    # Dias de trabalho distintos
    dias_trabalho = set()
    for _, row in df.iterrows():
        dias_trabalho.add((row["Professor"], row["Dia"]))
    total_dias_trabalho = len(dias_trabalho)

    # Janelas
    janelas = 0
    for prof in df["Professor"].unique():
        for dia in df["Dia"].unique():
            periodos = sorted(df[(df["Professor"] == prof) & (df["Dia"] == dia)]["Período"].tolist())
            if len(periodos) <= 1:
                continue
            for i in range(len(periodos) - 1):
                if periodos[i+1] - periodos[i] > 1:
                    janelas += periodos[i+1] - periodos[i] - 1

    ultima_linha = ws.max_row + 2

    ws.cell(row=ultima_linha, column=1, value="Dias de trabalho distintos")
    ws.cell(row=ultima_linha, column=2, value=total_dias_trabalho)
    ws.cell(row=ultima_linha+1, column=1, value="Total de janelas")
    ws.cell(row=ultima_linha+1, column=2, value=janelas)
    ws.cell(row=ultima_linha+2, column=1, value="Aulas duplas não atendidas")
    ws.cell(row=ultima_linha+2, column=2, value=duplas_nao_atendidas)

    wb.save("solucao_grade_horaria.xlsx")

if __name__ == "__main__":
    solucao_grade_horaria()
    calcular_metricas_e_adicionar_excel()


